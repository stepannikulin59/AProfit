import openpyxl

wb = openpyxl.load_workbook('test.xlsx')

print(wb.sheetnames)
sheet = wb.active

cell = sheet['A1']
print(cell)
print('Строка: ' + str(cell.row))
print('Столбец: ' + str(cell.column))
print('Ячейка: ' + str(cell.coordinate))
print('Значение: ' + str(cell.value))

# cell = sheet.cell(row=1, column=1)
# print(cell)
# print(cell.value)